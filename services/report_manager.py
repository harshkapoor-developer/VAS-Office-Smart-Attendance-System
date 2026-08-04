"""
services/report_manager.py
------------------------------
Reads attendance across one or more daily CSV files (via AttendanceManager)
and provides date-range queries, filtering, and export to CSV/Excel/PDF.

Has zero dependency on face_recognition/dlib - it's a pure read/aggregate/
export layer, fully testable without dlib installed.
"""

from __future__ import annotations

import csv
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import config
from services.attendance_manager import AttendanceManager
from utils.exceptions import ValidationError
from utils.logger import get_logger

logger = get_logger(__name__)


class ReportManager:
    def __init__(self, attendance_manager: Optional[AttendanceManager] = None) -> None:
        self.att_mgr = attendance_manager or AttendanceManager()

    # ------------------------------------------------------------------
    # Date-range queries
    # ------------------------------------------------------------------
    def read_range(self, start_date: date, end_date: date) -> list[dict]:
        """Reads and concatenates attendance rows across every day's CSV
        in [start_date, end_date] inclusive. Missing days (no one was
        marked, or the file doesn't exist yet) are silently skipped -
        that's a normal, expected gap, not an error.
        """
        if start_date > end_date:
            raise ValidationError("start_date must not be after end_date.")

        rows: list[dict] = []
        current = start_date
        while current <= end_date:
            day_rows = self.att_mgr.read_attendance_for_date(datetime(current.year, current.month, current.day))
            rows.extend(day_rows)
            current += timedelta(days=1)
        return rows

    def today(self) -> list[dict]:
        return self.att_mgr.read_today_attendance()

    def last_n_days(self, n: int) -> list[dict]:
        end = date.today()
        start = end - timedelta(days=n - 1)
        return self.read_range(start, end)

    def this_week(self) -> list[dict]:
        """Monday-to-today of the current week."""
        today = date.today()
        start = today - timedelta(days=today.weekday())
        return self.read_range(start, today)

    def this_month(self) -> list[dict]:
        today = date.today()
        start = today.replace(day=1)
        return self.read_range(start, today)

    # ------------------------------------------------------------------
    # Filtering
    # ------------------------------------------------------------------
    @staticmethod
    def filter_by_employee_search(rows: list[dict], search_term: str) -> list[dict]:
        term = search_term.lower()
        return [
            r for r in rows
            if term in r.get("Employee Name", "").lower() or term in r.get("Employee ID", "").lower()
        ]

    # ------------------------------------------------------------------
    # Summary stats (for dashboard charts / reports header)
    # ------------------------------------------------------------------
    @staticmethod
    def count_by_date(rows: list[dict]) -> dict[str, int]:
        return dict(Counter(r.get("Date", "Unknown") for r in rows))

    @staticmethod
    def count_by_employee(rows: list[dict]) -> dict[str, int]:
        return dict(Counter(r.get("Employee Name", "Unknown") for r in rows))

    # ------------------------------------------------------------------
    # Individual employee history: Date | In-Time | Out-Time | Status,
    # plus Total Working Days / Present / Absent / Attendance %.
    # ------------------------------------------------------------------
    def get_employee_history(
        self, employee_id: str, start_date: Optional[date] = None, end_date: Optional[date] = None
    ) -> dict:
        employee = self.att_mgr.emp_mgr.get_employee(employee_id)
        if employee is None:
            raise ValidationError(f"Employee ID '{employee_id}' does not exist.")

        end_date = end_date or date.today()
        if start_date is None:
            try:
                start_date = datetime.strptime(employee.joining_date, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                start_date = end_date
        if start_date > end_date:
            raise ValidationError("start_date must not be after end_date.")

        history: list[dict] = []
        present_days = 0
        current = start_date
        while current <= end_date:
            day_rows = self.att_mgr.read_attendance_for_date(datetime(current.year, current.month, current.day))
            match = next((r for r in day_rows if r.get("Employee ID") == employee_id), None)
            if match:
                history.append(match)
                present_days += 1
            else:
                history.append({
                    "Date": current.strftime(config.ATTENDANCE_DATE_DISPLAY_FORMAT),
                    "Employee Name": employee.name,
                    "Employee ID": employee_id,
                    "In-Time": "--",
                    "Out-Time": "--",
                    "Status": config.ATTENDANCE_STATUS_ABSENT,
                })
            current += timedelta(days=1)

        total_days = len(history)
        absent_days = total_days - present_days
        percentage = round((present_days / total_days) * 100, 1) if total_days else 0.0

        return {
            "employee": employee,
            "history": history,
            "total_working_days": total_days,
            "days_present": present_days,
            "days_absent": absent_days,
            "attendance_percentage": percentage,
        }

    def find_employees(self, search_term: str):
        """Search by name or ID for the employee-history lookup UI."""
        return self.att_mgr.emp_mgr.search_employees(search_term=search_term)

    # ------------------------------------------------------------------
    # Export: CSV
    # ------------------------------------------------------------------
    def export_csv(self, rows: list[dict], output_path: Path) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=config.ATTENDANCE_CSV_COLUMNS)
            writer.writeheader()
            for row in rows:
                writer.writerow({col: row.get(col, "") for col in config.ATTENDANCE_CSV_COLUMNS})
        logger.info("Exported %d row(s) to CSV: %s", len(rows), output_path)
        return output_path

    # ------------------------------------------------------------------
    # Export: Excel
    # ------------------------------------------------------------------
    def export_excel(self, rows: list[dict], output_path: Path) -> Path:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, col_name in enumerate(config.ATTENDANCE_CSV_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(config.ATTENDANCE_CSV_COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

        for col_idx, col_name in enumerate(config.ATTENDANCE_CSV_COLUMNS, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = max(14, len(col_name) + 2)

        wb.save(str(output_path))
        logger.info("Exported %d row(s) to Excel: %s", len(rows), output_path)
        return output_path

    # ------------------------------------------------------------------
    # Export: PDF
    # ------------------------------------------------------------------
    def export_pdf(self, rows: list[dict], output_path: Path, title: str = "Attendance Report") -> Path:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        output_path.parent.mkdir(parents=True, exist_ok=True)

        doc = SimpleDocTemplate(str(output_path), pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = [Paragraph(title, styles["Title"]), Spacer(1, 12)]

        table_data = [config.ATTENDANCE_CSV_COLUMNS] + [
            [row.get(col, "") for col in config.ATTENDANCE_CSV_COLUMNS] for row in rows
        ]
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ]))
        elements.append(table)

        doc.build(elements)
        logger.info("Exported %d row(s) to PDF: %s", len(rows), output_path)
        return output_path
